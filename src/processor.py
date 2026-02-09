import os
import re
import docx
import openpyxl
from utils import convert_to_mm  # Mantendo seu import original

# ==============================================================================
# 1. FUNÇÕES AUXILIARES E REGEX (NOVAS E ATUALIZADAS)
# ==============================================================================

def normalizar_nome_perfil_w(desc):
    """
    Padroniza o nome do perfil W para busca no Excel.
    Entrada: "W 150 x 22.5" ou "W150 22,5"
    Saída: "W150X22,5" (Padrão comum em catálogos brasileiros)
    """
    # Regex explicaçao:
    # W\s*       -> Letra W seguida ou não de espaço
    # (\d+)      -> Captura a Altura (primeiro número)
    # [xX\s]*    -> Separador (letra x, X ou apenas espaço)
    # ([\d\.,]+) -> Captura o Peso/Bitola (número, ponto ou vírgula)
    match = re.search(r'W\s*(\d+)\s*[xX\s]*\s*([\d\.,]+)', desc, re.IGNORECASE)
    
    if match:
        altura = match.group(1)
        peso = match.group(2).replace('.', ',') # Força vírgula para bater com strings PT-BR do Excel
        
        # Remove ,0 se for inteiro (ex: 18,0 -> 18) para evitar erro de busca
        if peso.endswith(',0'): 
            peso = peso[:-2]
            
        return f"W{altura}X{peso}"
    return desc.upper().strip()

def extrair_comprimento_texto(texto):
    """
    Tenta extrair o comprimento de dentro da string de descrição se não houver na coluna.
    Procura por: C=5000, L:5000mm, 5000mm, etc.
    Retorna valor em METROS (float).
    """
    # Procura padrões explícitos: C=..., L=..., ou número seguido de 'mm'
    # Ex: "C=4500", "L: 3000", "4500mm"
    regex_mm = r'(?:C|L|COMPR)[\s=:]*(\d{3,5})|(?:^|\s)(\d{3,5})\s*mm'
    match = re.search(regex_mm, texto, re.IGNORECASE)
    
    if match:
        # Pega o primeiro grupo que não for None
        valor_mm_str = match.group(1) if match.group(1) else match.group(2)
        try:
            return float(valor_mm_str) / 1000.0  # Converte mm -> m
        except ValueError:
            return 0.0
    return 0.0

def classificar_e_mapear_perfil(desc):
    """Identifica o TIPO de perfil e retorna o código e uma chave de classificação."""
    desc_upper = desc.upper()
    
    # Prioridade para VIGA W detectada via Regex robusto
    if re.search(r'W\s*\d', desc_upper):
        return 'VIGA W', 'VIGA_W'

    if '[' in desc_upper or '][' in desc_upper: return 'U.s', 'PERFIL_U'
    if 'UENR' in desc_upper or 'IENR' in desc_upper or 'CART' in desc_upper or 'CA ' in desc_upper: return 'U.e', 'TERCA'
    if 'L DOBRADO' in desc_upper or desc_upper.startswith('L '): return 'L DOBRADO', 'CANTONEIRA'
    if 'RED' in desc_upper: return 'FERRO MECANICO RED.', 'TUBO' 
    if 'TUBO' in desc_upper: return 'TUBO', 'TUBO' 

    return 'N/D', 'OUTROS'

def parse_dimensoes_inteligente(desc, tipo_perfil):
    """Aplica regras de extração de dimensões e retorna as medidas."""
    a, b, c, esp = 0.0, 0.0, 0.0, 0.0
    # Extrai todos os números
    numeros_str_list = re.findall(r'[\d\./,]+', desc) # Adicionado vírgula no regex

    if tipo_perfil in ['PERFIL_U']:
        if len(numeros_str_list) >= 3:
            a = convert_to_mm(numeros_str_list[0])
            b = convert_to_mm(numeros_str_list[1])
            esp = convert_to_mm(numeros_str_list[2])
    elif tipo_perfil == 'TERCA':
        if len(numeros_str_list) >= 4:
            a = convert_to_mm(numeros_str_list[0])
            b = convert_to_mm(numeros_str_list[1])
            c = convert_to_mm(numeros_str_list[2])
            esp = convert_to_mm(numeros_str_list[3])
    elif tipo_perfil == 'CANTONEIRA':
        if len(numeros_str_list) >= 2:
            # Lógica simplificada
            aba = convert_to_mm(numeros_str_list[0])
            a, b = aba, aba
            if len(numeros_str_list) > 1:
                esp = convert_to_mm(numeros_str_list[1])
    elif tipo_perfil == 'TUBO':
        if len(numeros_str_list) >= 1:
            esp = convert_to_mm(numeros_str_list[0])
            
    # OBS: VIGA_W não precisa extrair dimensões aqui pois usaremos a busca por nome no Excel
    # ou podemos extrair apenas para log.
    
    return a, b, c, esp

# ==============================================================================
# 2. EXTRAÇÃO DE DADOS (DOCX)
# ==============================================================================

def extrair_dados_word(caminho_arquivo_word):
    """Lê a tabela do Word e processa linhas, incluindo lógica específica para W."""
    documento = docx.Document(caminho_arquivo_word)
    try:
        tabela = documento.tables[0]
    except IndexError:
        return None

    if len(tabela.rows) < 2: return None

    # Extração "Bruta" das colunas
    perfils_str = tabela.cell(1, 0).text
    acos_str = tabela.cell(1, 1).text
    ltotais_str = tabela.cell(1, 2).text
    pesos_str = tabela.cell(1, 3).text

    lista_perfis = list(filter(None, perfils_str.strip().split('\n')))
    lista_acos = list(filter(None, acos_str.strip().split('\n')))
    lista_ltotais = list(filter(None, ltotais_str.strip().split('\n')))
    lista_pesos = list(filter(None, pesos_str.strip().split('\n')))

    num_perfis = len(lista_perfis)
    # Validação básica de quantidades
    if num_perfis == 0: return None
    
    dados_finais = []
    
    for i in range(num_perfis):
        perfil = lista_perfis[i].strip()
        aco = lista_acos[i].strip() if i < len(lista_acos) else (lista_acos[0].strip() if lista_acos else "A36")
        
        # --- Lógica Híbrida de Comprimento ---
        l_total_m = 0.0
        
        # 1. Tenta pegar da coluna da tabela (Prioridade 1 se existir)
        val_coluna = lista_ltotais[i].strip() if i < len(lista_ltotais) else ""
        if val_coluna:
            try:
                # CORREÇÃO: Dividimos por 100 para converter de cm para m, pois o Excel pode ter valores em cm
                l_total_m = float(val_coluna.replace(',', '.')) / 100.0
            except ValueError:
                l_total_m = 0.0
        
        # 2. Se a coluna for 0 ou vazia, tenta extrair do texto da descrição (Prioridade 2)
        if l_total_m == 0.0:
            l_total_m = extrair_comprimento_texto(perfil)

        # Peso
        peso_str = lista_pesos[i].strip().replace(',', '.') if i < len(lista_pesos) else "0"
        try:
            peso_final = float(peso_str)
        except ValueError:
            peso_final = 0.0

        dados_finais.append([perfil, aco, l_total_m, peso_final])

    return dados_finais

# ==============================================================================
# 3. MANIPULAÇÃO DO EXCEL
# ==============================================================================

def encontrar_proxima_linha_vazia(sheet, codigo_secao, linha_inicio_busca):
    for row in range(linha_inicio_busca, sheet.max_row + 2):
        celula_codigo = sheet.cell(row=row, column=1)
        celula_dado_ref = sheet.cell(row=row, column=2)
        # Verifica se o código bate e se a coluna B (dado ref) está "disponível"
        # Adicionei checagem para 'X' ou None
        if str(celula_codigo.value).strip() == codigo_secao:
            # Se coluna J (comprimento) estiver vazia, considera linha livre
            if sheet.cell(row=row, column=10).value in [None, 0, '', '0', 0.0]:
                return row
    return None

def preencher_planilha_excel(caminho_planilha, dados_materiais):
    """
    Carrega o template, processa os dados (incluindo CA e W) e salva como '_processado'.
    """
    workbook = openpyxl.load_workbook(caminho_planilha)
    sheet = workbook.active
    
    # 1. GERAÇÃO DO NOVO NOME DE ARQUIVO (_processado)
    nome_base, extensao = os.path.splitext(caminho_planilha)
    caminho_processado = f"{nome_base}_processado{extensao}"
    
    # Agrupa dados para processar em ordem
    dados_agrupados = {}
    for item in dados_materiais:
        perfil_raw = item[0]
        codigo_excel, tipo_perfil = classificar_e_mapear_perfil(perfil_raw)
        chave = 'VIGA_W_GROUP' if tipo_perfil == 'VIGA_W' else codigo_excel
        if chave not in dados_agrupados: dados_agrupados[chave] = []
        dados_agrupados[chave].append(item)

    # 2. PROCESSAMENTO DOS ITENS
    for chave_grupo, itens_da_secao in dados_agrupados.items():
        linha_de_busca_secao = 4
        
        for item in itens_da_secao:
            perfil_desc, aco_tipo, l_total_m, peso_total = item
            _, tipo_perfil = classificar_e_mapear_perfil(perfil_desc)
            
            # Melhoria CA: Dobra a metragem
            if 'CA ' in perfil_desc.upper():
                l_total_m = l_total_m * 2
            
            linha_alvo = None
            
            # Identificação da linha (W por Nome, Outros por Próxima Vazia)
            if tipo_perfil == 'VIGA_W':
                nome_normalizado = normalizar_nome_perfil_w(perfil_desc)
                for r in range(4, sheet.max_row + 1):
                    val_cell = str(sheet.cell(row=r, column=1).value).strip().upper().replace(" ", "")
                    if val_cell == nome_normalizado.replace(" ", ""):
                        if sheet.cell(row=r, column=10).value in [None, 0, '', 0.0]:
                            linha_alvo = r
                            break
            else:
                linha_alvo = encontrar_proxima_linha_vazia(sheet, chave_grupo, linha_de_busca_secao)

            # --- GRAVAÇÃO DOS DADOS (CORREÇÃO AQUI) ---
            if linha_alvo:
                dim_a, dim_b, dim_c, dim_esp = parse_dimensoes_inteligente(perfil_desc, tipo_perfil)
                
                # Se NÃO for Viga W, preenchemos as dimensões técnicas (A, B, C, esp)
                if tipo_perfil != 'VIGA_W':
                    if tipo_perfil in ['PERFIL_U', 'TERCA']:
                        sheet.cell(row=linha_alvo, column=2).value = dim_a
                        sheet.cell(row=linha_alvo, column=4).value = dim_b
                        sheet.cell(row=linha_alvo, column=6).value = dim_c
                    elif tipo_perfil == 'CANTONEIRA':
                        sheet.cell(row=linha_alvo, column=4).value = dim_a
                        sheet.cell(row=linha_alvo, column=6).value = dim_b
                    
                    # Espessura sempre na coluna 8 para perfis dobrados/tubos
                    sheet.cell(row=linha_alvo, column=8).value = dim_esp

                # Colunas fixas para todos os perfis
                sheet.cell(row=linha_alvo, column=9).value = aco_tipo
                if l_total_m > 0:
                    sheet.cell(row=linha_alvo, column=10).value = l_total_m
                sheet.cell(row=linha_alvo, column=17).value = peso_total
                
                # Atualiza o ponteiro de busca para não sobrescrever a mesma linha
                if tipo_perfil != 'VIGA_W':
                    linha_de_busca_secao = linha_alvo + 1

    # 3. LIMPEZA VISUAL E SALVAMENTO
    ocultar_linhas_vazias(sheet)
    
    workbook.save(caminho_processado)
    print(f"\n[OK] Processamento finalizado com sucesso!")
    print(f"[ARQUIVO] Gerado: {caminho_processado}")

def ocultar_linhas_vazias(sheet, linha_inicio=4):
    """
    Oculta linhas sem metragem, protegendo o bloco de resumo.
    """
    for row in range(linha_inicio, sheet.max_row + 1):
        valor_A = str(sheet.cell(row=row, column=1).value).strip().upper()
        
        # Trava para manter o resumo visível (TOTAL ou ATIVO FINAL)
        if "TOTAL" in valor_A or "ATIVO FINAL" in valor_A or "RESUMO" in valor_A:
            for r_resumo in range(row, min(row + 25, sheet.max_row + 1)):
                sheet.row_dimensions[r_resumo].hidden = False
            break 
            
        valor_comprimento = sheet.cell(row=row, column=10).value
        if sheet.cell(row=row, column=1).value and (valor_comprimento in [None, 0, '0', '', 0.0]):
            sheet.row_dimensions[row].hidden = True
        else:
            sheet.row_dimensions[row].hidden = False